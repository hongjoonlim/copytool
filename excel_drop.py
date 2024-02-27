import tkinter as tk
from tkinter import filedialog
from tkinterdnd2 import DND_FILES, TkinterDnD
from openpyxl import load_workbook

class DragDropApp:
    def __init__(self):
        self.root = TkinterDnD.Tk()
        self.root.title("Drag and Drop Demo")
        self.root.geometry("600x400")

        self.left_frame = tk.Frame(self.root, width=300, height=400, bg="lightgray")
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.right_frame = tk.Frame(self.root, width=300, height=400, bg="lightblue")
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.left_label = tk.Label(self.left_frame, text="Drop Files Here", bg="lightgray")
        self.left_label.pack(pady=10)

        self.text_widget = tk.Text(self.left_frame, wrap=tk.WORD)
        self.text_widget.pack(fill=tk.BOTH, expand=True)

        # Bind drop event to handle files dropped onto the left frame
        self.left_frame.drop_target_register(DND_FILES)
        self.left_frame.dnd_bind('<<Drop>>', self.drop)

        self.copy_button = tk.Button(self.right_frame, text="Copy", command=self.copy_folder_name)
        self.copy_button.pack(pady=20)

        self.select_dest_button = tk.Button(self.right_frame, text="Select Destination Folder", command=self.select_destination_path)
        self.select_dest_button.pack(pady=10)

        self.destination_path_label = tk.Label(self.right_frame, text="", bg="lightblue")
        self.destination_path_label.pack()

        self.destination_path = ""

    def drop(self, event):
        self.left_label.config(text="Dropped")
        file_path = event.data.strip()  # Get dropped file path
        if file_path.lower().endswith(('.xlsx', '.xls')):  # Check if it's an Excel file
            # Load the Excel file and extract the selected cell values
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                selected_cells = []
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.coordinate in ws._cells_by_col:
                            selected_cells.append(cell.value)
                self.text_widget.delete(1.0, tk.END)
                self.text_widget.insert(tk.END, '\n'.join(str(cell) for cell in selected_cells))
            except Exception as e:
                print("Error:", e)
        else:
            print("Unsupported file format.")

    def copy_folder_name(self):
        if self.destination_path:
            print("Copying folder name to destination path:", self.destination_path)
            # Here you can implement the logic to copy folder name to the destination path
        else:
            print("Please select a destination path.")

    def select_destination_path(self):
        self.destination_path = filedialog.askdirectory()
        self.destination_path_label.config(text="Destination Path: " + self.destination_path)

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = DragDropApp()
    app.run()